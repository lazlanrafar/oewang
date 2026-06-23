"""Document text extraction + overlapping chunking for RAG indexing. Port of
packages/ai/embedding/chunking.service.ts. PDF via pypdf, XLSX via openpyxl, CSV/text
decoded directly. Elysia writes the resulting chunks+embeddings to vault_file_chunks.
"""

import base64 as b64
import csv
import io
import re

CHUNK_SIZE = 1000  # characters
CHUNK_OVERLAP = 200

_TEXT_TYPES = ("application/json", "application/xml", "text/csv")
_XLSX = (
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)


def estimate_tokens(text: str) -> int:
    return -(-len(text) // 4)  # ceil(len/4)


def is_indexable(mime_type: str) -> bool:
    t = mime_type.lower()
    return (
        t.startswith("text/")
        or t == "application/pdf"
        or t in ("application/json", "application/xml")
        or t in _XLSX
    )


def chunk(text: str) -> list[dict]:
    """Overlapping ~1000-char chunks, broken on word boundaries. Mirrors the TS
    pointer arithmetic exactly (incl. the tiny-text infinite-loop guard)."""
    cleaned = re.sub(r"\n{3,}", "\n\n", text.replace("\r\n", "\n")).strip()
    if not cleaned:
        return []

    chunks: list[dict] = []
    pos, index = 0, 0
    n = len(cleaned)
    while pos < n:
        end = min(pos + CHUNK_SIZE, n)
        piece = cleaned[pos:end]
        if end < n:
            break_at = max(piece.rfind(" "), piece.rfind("\n"))
            if break_at > CHUNK_SIZE / 2:
                piece = piece[:break_at]
        content = piece.strip()
        if content:
            chunks.append({"content": content, "index": index, "tokenCount": estimate_tokens(content)})
            index += 1
        # Advance by (chunk - overlap). If that's <= 0 the remaining tail is already
        # covered by the previous chunk's overlap, so stop — guards the infinite loop
        # the TS original has when the final piece length == CHUNK_OVERLAP.
        step = len(piece) - CHUNK_OVERLAP
        if step <= 0:
            break
        pos += step
    return chunks


def extract_text(data_b64: str, mime_type: str) -> str | None:
    """Plain text / PDF / XLSX / CSV → text. None for unsupported (images, binary)."""
    buf = b64.b64decode(data_b64)
    t = mime_type.lower()

    if t.startswith("text/") or t in ("application/json", "application/xml"):
        return buf.decode("utf-8", errors="replace")

    if t == "application/pdf":
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(buf))
            return "\n\n".join((p.extract_text() or "") for p in reader.pages).strip()
        except Exception:  # noqa: BLE001
            return None

    if t in _XLSX:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
            sheets = []
            for ws in wb.worksheets:
                out = io.StringIO()
                w = csv.writer(out)
                for row in ws.iter_rows(values_only=True):
                    w.writerow(["" if c is None else c for c in row])
                sheets.append(f"[Sheet: {ws.title}]\n{out.getvalue()}")
            return "\n\n".join(sheets).strip()
        except Exception:  # noqa: BLE001
            return None

    return None
