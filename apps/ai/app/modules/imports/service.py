"""Extract transactions from tabular CSV/XLSX rows — OpenAI-only port of
packages/ai/core/extraction.service.ts. Elysia still writes the rows; this only
classifies them.
"""

import base64 as b64
import csv
import io
import json
import logging

from app.config import get_settings
from app.core.llm import get_client

log = logging.getLogger("ai.imports")


def parse_file_to_rows(data_b64: str, mime_type: str) -> list[dict]:
    """CSV/XLSX bytes → list of header→value dicts. Port of parseFileToAiInput
    (the tabular branch). Non-tabular types yield []."""
    buf = b64.b64decode(data_b64)
    t = mime_type.lower()

    if t == "text/csv" or t.startswith("text/"):
        text = buf.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(r) for r in reader]

    if t in (
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(next(rows_iter))]
        except StopIteration:
            return []
        out = []
        for row in rows_iter:
            out.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
        return out

    return []

_TX_PROPS = {
    "name": {"type": "string"},
    "amount": {"type": "number"},
    "date": {"type": "string"},
    "type": {"type": "string", "enum": ["income", "expense", "transfer"]},
    "walletName": {"type": ["string", "null"]},
    "categoryName": {"type": ["string", "null"]},
    "description": {"type": ["string", "null"]},
}

_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "transactions": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": _TX_PROPS,
                "required": list(_TX_PROPS.keys()),
            },
        }
    },
    "required": ["transactions"],
}


def extract_transactions(
    rows: list[dict], wallet_names: list[str], category_names: list[str]
) -> list[dict]:
    if not rows or not get_settings().OPENAI_API_KEY:
        return []

    wallets = "\n".join(f"- {w}" for w in wallet_names)
    cats = "\n".join(f"- {c}" for c in category_names)
    prompt = f"""You are a financial data extraction tool.
Extract transactions from the following tabular data (CSV/Excel rows).

Valid Wallets (Account Names):
{wallets}

Valid Categories:
{cats}

Rules:
- Extract up to 100 rows.
- Match walletName against the provided wallet list whenever possible.
- Match categoryName against the provided category list whenever possible.
- Use ISO 8601 dates.
- Ignore rows that are clearly headers, totals, separators, or blank.

Tabular Data:
{json.dumps(rows[:100])}""".strip()

    try:
        resp = get_client().chat.completions.create(
            model=get_settings().AI_CHAT_MODEL,
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}],
            response_format={
                "type": "json_schema",
                "json_schema": {"name": "extracted_transactions", "strict": True, "schema": _SCHEMA},
            },
        )
        parsed = json.loads(resp.choices[0].message.content or '{"transactions":[]}')
        txns = parsed.get("transactions")
        return txns if isinstance(txns, list) else []
    except Exception as e:  # noqa: BLE001
        log.error("extract_transactions failed: %s", e)
        return []
